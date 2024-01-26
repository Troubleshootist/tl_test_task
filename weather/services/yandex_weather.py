from dataclasses import dataclass, asdict
from datetime import datetime
from openpyxl import load_workbook

import requests

from weather.exceptions import (
    YandexWeatherDataFetchError,
    FileWithCoordsNotFound,
    CoordsParsingException,
    NoCityFound,
)
from weather_project_tl.settings import (
    YANDEX_WEATHER_URL,
    YANDEX_WEATHER_API_KEY,
    BASE_DIR,
)


@dataclass
class Weather:
    temperature: int
    pressure: int
    wind_speed: int
    observed_at: datetime

    def dict(self):
        return asdict(self)


def get_lat_and_lon(city_name: str) -> (float, float):
    try:
        xlsx_file_path = BASE_DIR / "weather" / "services" / "koord_lat_lng_russia.xlsx"
        workbook = load_workbook(filename=xlsx_file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=True):
            if city_name.lower() in row[0].lower():
                lat, lon = row[3], row[4]
                workbook.close()
                return lat, lon
        workbook.close()

    except FileNotFoundError as e:
        raise FileWithCoordsNotFound from e
    except Exception as e:
        raise CoordsParsingException from e
    raise NoCityFound()


def get_yandex_weather(city_name: str) -> Weather:
    lat, lon = get_lat_and_lon(city_name)
    params = {
        "lat": lat,
        "lon": lon,
    }
    headers = {"X-Yandex-API-Key": YANDEX_WEATHER_API_KEY}

    try:
        response = requests.get(url=YANDEX_WEATHER_URL, params=params, headers=headers)
        response.raise_for_status()

        current_weather_data = response.json().get("fact")

    except requests.exceptions.RequestException as e:
        raise YandexWeatherDataFetchError from e

    return Weather(
        temperature=current_weather_data.get("temp"),
        pressure=current_weather_data.get("pressure_mm"),
        wind_speed=current_weather_data.get("wind_speed"),
        observed_at=datetime.fromtimestamp(current_weather_data.get("obs_time")),
    )
